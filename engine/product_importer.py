# -*- coding: utf-8 -*-
"""LiuGong Product Importer v4.0 - Extracts: model, specs, DAP price, scrap tax"""
import os, json, re
from datetime import datetime
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PRODUCTS_DIR = os.path.join(BASE_DIR, 'products')
DB_PATH = os.path.join(BASE_DIR, 'engine', 'product_db.json')

CATEGORY_RULES = [
    (['WL', '装载机', 'loader'], '装载机'),
    (['EX', '挖掘机', '挖机', 'excavator'], '挖掘机'),
    (['RL', '压路机', 'roller', '振动'], '压路机'),
    (['BD', '推土机', 'bulldozer'], '推土机'),
    (['MG', '平地机', 'grader'], '平地机'),
    (['MT', '矿卡', 'mining truck'], '宽体矿车'),
    (['SSL', '滑移', 'skid', 'steer'], '滑移装载机'),
    (['BHL', '两头忙', 'backhoe'], '两头忙'),
    (['Electric', '电动', 'electric'], '电动设备'),
    (['AWP', 'awp', '高空'], '高空作业平台'),
    (['Crane', '起重机', 'crane', 'TC'], '汽车起重机'),
    (['Truck Crane', '汽车吊'], '汽车起重机'),
]

def detect_file_category(fname):
    fn = str(fname).lower()
    for kws, cat in CATEGORY_RULES:
        if any(kw.lower() in fn for kw in kws):
            return cat
    return None

def extract_tonnage(model_str):
    m = re.search(r'(\d{2,3})', str(model_str))
    if not m: return None
    v = int(m.group(1))
    return v // 10 if v >= 100 else v

def load_scrap_tax(filepath):
    """Load scrap tax data: model -> {hp, tax_2026_rub}"""
    if not os.path.exists(filepath):
        return {}
    wb = openpyxl.load_workbook(filepath, data_only=True)
    tax_map = {}
    for sn in ['Latest version', 'Old version']:
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        # Find header
        model_col, hp_col, tax_col = None, None, None
        for r in range(1, min(5, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                h = str(ws.cell(row=r, column=c).value or '').strip()
                if '机型' in h or h == 'Model':
                    model_col = c
                if h and str(h).strip() in ['马力', 'HP']:
                    hp_col = c
                if '26年报废税' in h or 'Tax 2026' in h:
                    tax_col = c
        if not model_col or not tax_col: continue
        for r in range(2, ws.max_row + 1):
            model = str(ws.cell(row=r, column=model_col).value or '').strip()
            if not model or model == 'None': continue
            hp = ws.cell(row=r, column=hp_col).value if hp_col else None
            tax = ws.cell(row=r, column=tax_col).value
            if tax is not None:
                try: tax = float(tax)
                except: tax = 0
            tax_map[model.upper().replace(' ', '')] = {
                'scrap_tax_rub': tax if tax else 0,
                'hp': float(str(hp)) if hp and str(hp).replace('.','').isdigit() else 0,
                'scrap_tax_year': 2026,
            }
    wb.close()
    return tax_map

def parse_price_list_sheet(ws, sn):
    """Extract model info + FOB + DAP price from a price list sheet"""
    product = {'model': sn.strip(), 'source_sheet': sn}

    fob_price = None
    dap_price = None
    in_fob = False
    in_dap = False

    for r in range(1, min(ws.max_row + 1, 40)):
        for c in range(1, min(ws.max_column + 1, 6)):
            val = ws.cell(row=r, column=c).value
            if not val: continue
            vs = str(val).strip()

            # Detect section headers
            if 'FOB' in vs.upper() and 'PRICE' in vs.upper():
                in_fob, in_dap = True, False
                continue
            if 'DAP' in vs.upper() and ('PRICE' in vs.upper() or '满洲里' in vs):
                in_fob, in_dap = False, True
                continue

            # Extract model name from description
            if 'CLG' in vs.upper() and len(vs) > 30:
                product['desc'] = vs[:250]
                m = re.search(r'(?:康明斯|潍柴|玉柴|云内|东康|珀金斯)[^\s/,]*', vs)
                if m: product['engine'] = m.group(0)
                m = re.search(r'(?:Stage\s*)?(I{2,3}[A-Z]?|欧[IV]+|[二三四五]\s*阶段)', vs, re.I)
                if m: product['emission'] = m.group(0)

            # Extract price
            if isinstance(val, (int, float)) and val > 100:
                if in_fob and not fob_price:
                    fob_price = float(val)
                elif in_dap and not dap_price:
                    dap_price = float(val)

    product['fob_price_cny'] = fob_price
    product['dap_price_cny'] = dap_price
    product['tonnage'] = extract_tonnage(sn)
    return product

def parse_excavator_remark(ws):
    """Parse excavator Remark sheet for model comparison table"""
    products = []
    # Find the comparison table
    for r in range(1, min(ws.max_row + 1, 120)):
        model_val = ws.cell(row=r, column=2).value
        if not model_val: continue
        model_str = str(model_val).strip()
        # Skip non-model rows
        if not re.search(r'\d', model_str) or 'CLG' not in model_str.upper():
            continue
        if 'LGCE' in model_str.upper() or len(model_str) < 4:
            continue

        desc = ws.cell(row=r, column=3).value
        fob = ws.cell(row=r, column=5).value
        dap = ws.cell(row=r, column=6).value

        product = {
            'model': model_str,
            'source_sheet': 'Remark',
        }
        if desc: product['desc'] = str(desc).strip()[:200]
        if fob:
            try: product['fob_price_cny'] = float(fob)
            except: pass
        if dap:
            try: product['dap_price_cny'] = float(dap)
            except: pass

        m = re.search(r'(?:康明斯|潍柴|珀金斯)[^\s/,]*', str(desc or ''))
        if m: product['engine'] = m.group(0)
        product['tonnage'] = extract_tonnage(model_str)
        products.append(product)
    return products

def import_all():
    if not os.path.exists(PRODUCTS_DIR):
        print('[ERROR] products/ folder not found'); return None

    xlsx_files = [f for f in os.listdir(PRODUCTS_DIR) if f.endswith(('.xlsx','.xls')) and not f.startswith('~')]

    # Load scrap tax
    scrap_file = os.path.join(PRODUCTS_DIR, '2026年报废税 2026Y scrap tax.xlsx')
    scrap_tax = load_scrap_tax(scrap_file)
    print('Scrap tax loaded: {} models'.format(len(scrap_tax)))

    all_products = []
    print('\nProcessing {} files:'.format(len(xlsx_files)))

    for fname in sorted(xlsx_files):
        if '报废税' in fname or 'scrap tax' in fname.lower():
            continue
        fpath = os.path.join(PRODUCTS_DIR, fname)
        cat = detect_file_category(fname)
        if not cat:
            print('  [SKIP] {} - unknown category'.format(fname[:50]))
            continue

        try:
            if fname.endswith('.xls'):
                continue  # skip .xls for now
            wb = openpyxl.load_workbook(fpath, data_only=True)
            sheets = wb.sheetnames
            count = 0

            # Special handling for excavator Remark sheet
            if any('remark' in s.lower() for s in sheets):
                ws_remark = wb[[s for s in sheets if 'remark' in s.lower()][0]]
                products = parse_excavator_remark(ws_remark)
                for p in products:
                    p['category'] = cat
                    p['source_file'] = fname
                    # Look up scrap tax
                    key = p['model'].upper().replace(' ', '')
                    if key in scrap_tax:
                        p.update(scrap_tax[key])
                all_products.extend(products)
                count += len(products)

            # Process individual model sheets
            model_sheets = [s for s in sheets if s.lower() not in ['remark', 'catalogue', 'sheet1'] and re.search(r'\d', s)]
            for sn in model_sheets:
                ws = wb[sn]
                p = parse_price_list_sheet(ws, sn)
                p['category'] = cat
                p['source_file'] = fname

                # Match with scrap tax by model name
                model_key = sn.strip().upper().replace(' ', '')
                if model_key in scrap_tax:
                    p.update(scrap_tax[model_key])

                # Also match CLGxxx variant
                if p.get('desc'):
                    m = re.search(r'(CLG\s*\d+[A-Za-z]*)', p['desc'].upper())
                    if m:
                        clg_key = m.group(1).replace(' ', '')
                        if clg_key in scrap_tax and not p.get('scrap_tax_rub'):
                            p.update(scrap_tax[clg_key])

                all_products.append(p)
                count += 1

            wb.close()
            print('  [{}] {} -> {} products'.format(cat, fname[:40], count))
        except Exception as e:
            print('  [ERROR] {}: {}'.format(fname[:40], e))

    # Deduplicate
    seen = {}
    for p in all_products:
        key = p['model'].upper().replace(' ', '').replace('-', '')
        if key not in seen or len(str(p.get('desc', ''))) > len(str(seen[key].get('desc', ''))):
            # Prefer entries with DAP price
            if key in seen and p.get('dap_price_cny') and not seen[key].get('dap_price_cny'):
                seen[key] = p
            elif key not in seen:
                seen[key] = p

    unique = list(seen.values())
    unique = [p for p in unique if re.search(r'\d', p['model']) and len(p['model']) >= 3]

    # Stats
    from collections import Counter
    cats = Counter(p.get('category', '?') for p in unique)
    with_price = sum(1 for p in unique if p.get('dap_price_cny'))
    with_tax = sum(1 for p in unique if p.get('scrap_tax_rub'))

    print('\n=== Import Summary ===')
    print('Total unique products: {}'.format(len(unique)))
    print('With DAP price: {}'.format(with_price))
    print('With scrap tax: {}'.format(with_tax))
    print('Categories:')
    for c, n in cats.most_common():
        print('  {}: {}'.format(c, n))

    db = {
        'meta': {
            'version': datetime.now().strftime('%Y%m%d-%H%M%S'),
            'total': len(unique),
            'with_price': with_price,
            'with_tax': with_tax,
            'files': xlsx_files,
        },
        'products': unique,
        'scrap_tax_base': 172500,
        'scrap_tax_year': 2026,
    }
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with open(DB_PATH, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
    print('Database saved: ' + DB_PATH)
    return db

if __name__ == '__main__':
    import_all()
