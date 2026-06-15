
import os, json, re, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PRODUCTS_DIR = os.path.join(os.path.dirname(BASE_DIR), 'products')
OUTPUT_PATH = os.path.join(BASE_DIR, 'product_db_full.json')

def safe_str(v):
    if v is None: return ''
    return str(v).strip()

def find_price(row, start=2):
    for i in range(start, min(len(row), 10)):
        v = row[i]
        if v is None: continue
        try:
            fv = float(v)
            if 1000 < fv < 50000000: return fv
        except: pass
    return None

def determine_category(filename):
    fn = filename.upper()
    if 'EX' in fn: return 'EX'
    if 'WL' in fn: return 'WL'
    if 'RL' in fn: return 'RL'
    if 'BD' in fn: return 'BD'
    if 'MG' in fn: return 'MG'
    if 'SSL' in fn: return 'SSL'
    if 'BHL' in fn: return 'BHL'
    if 'MT' in fn: return 'MT'
    if 'AWP' in fn: return 'AWP'
    if 'TOWER' in fn: return 'TOWER'
    if 'TRUCK' in fn: return 'TRUCK'
    if 'DRILL' in fn: return 'DRILL'
    if 'AIR' in fn: return 'AIR'
    if 'TELEHANDLER' in fn: return 'TELEHANDLER'
    if 'ELECTRIC' in fn: return 'ELECTRIC'
    return 'OTHER'

def extract_products(filepath, filename):
    products = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except:
        print(f'  SKIP {filename}')
        return products
    
    skip = {'remark','catalogue','skidsteer loader','backhoe loader','compact wheel loader','lgru-russia-rl'}
    

    # Check for RL consolidated sheet format (LGRU-RUSSIA-RL)
    rl_sheet = None
    for sn in wb.sheetnames:
        snu = sn.upper().replace(' ', '').replace('-', '')
        if 'LGRURUSSIARL' in snu:
            rl_sheet = sn
            break
    
    if rl_sheet:
        ws = wb[rl_sheet]
        rows_list = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        # Find header row with FOB and DAP
        hdr_row = -1
        for i, row in enumerate(rows_list):
            rt = ' '.join([safe_str(c) for c in row])
            if 'FOB' in rt.upper() and 'DAP' in rt.upper():
                hdr_row = i
                break
        if hdr_row >= 0:
            hdr = [safe_str(c) for c in rows_list[hdr_row]]
            model_col = fob_col = dap_col = -1
            for j, h in enumerate(hdr):
                hu = h.upper()
                if 'MODEL' in hu or '型号' in h:
                    model_col = j
                if 'FOB' in hu:
                    fob_col = j
                if 'DAP' in hu:
                    dap_col = j
            
            for i in range(hdr_row + 1, len(rows_list)):
                row = rows_list[i]
                model = safe_str(row[model_col]) if model_col >= 0 and model_col < len(row) else ''
                if not model or len(model) < 3:
                    continue
                fob_rl = None
                dap_rl = None
                try:
                    if fob_col >= 0 and fob_col < len(row):
                        fob_rl = float(row[fob_col]) if row[fob_col] else None
                except:
                    pass
                try:
                    if dap_col >= 0 and dap_col < len(row):
                        dap_rl = float(row[dap_col]) if row[dap_col] else None
                except:
                    pass
                
                if fob_rl and fob_rl > 1000:
                    desc_en_rl = safe_str(row[4]) if len(row) > 4 else ''
                    desc_cn_rl = safe_str(row[3]) if len(row) > 3 else ''
                    eng, hp = parse_eng_hp(desc_en_rl + ' ' + desc_cn_rl)
                    products.append({
                        'model': model,
                        'category': determine_category(filename, sn),
                        'desc_ru': desc_en_rl[:300],
                        'desc_cn': desc_cn_rl[:300],
                        'fob_price_cny': fob_rl,
                        'dap_price_cny': dap_rl if dap_rl else fob_rl,
                        'engine': eng,
                        'hp': hp,
                        'emission': '',
                        'bucket_m3': 0,
                        'tonnage': 0,
                        'source_file': filename,
                        'source_sheet': rl_sheet,
                    })
        wb.close()
        return products

    for sn in wb.sheetnames:
        if sn.lower().strip() in skip: continue
        if sn in ['Catalogue','Remark','SkidSteer Loader','Backhoe Loader']: continue
        try:
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=True))
            fob = None
            dap = None
            desc_en = ''
            desc_cn = ''
            emission = ''
            
            for i, row in enumerate(rows):
                rv = [safe_str(c) for c in row]
                rt = ' '.join(rv)
                
                if 'FOB' in rt and 'Price' in rt:
                    for j in range(i+1, min(i+4, len(rows))):
                        p = find_price(rows[j], 1)
                        if p:
                            fob = p
                            desc_en = safe_str(rows[j][0]) if rows[j][0] else ''
                            desc_cn = safe_str(rows[j][1]) if len(rows[j]) > 1 else ''
                            break
                
                if 'DAP' in rt and ('Price' in rt or 'Manzhouli' in rt):
                    for j in range(i+1, min(i+4, len(rows))):
                        p = find_price(rows[j], 1)
                        if p:
                            dap = p
                            if not desc_en:
                                desc_en = safe_str(rows[j][0]) if rows[j][0] else ''
                                desc_cn = safe_str(rows[j][1]) if len(rows[j]) > 1 else ''
                            break
                
                if 'Stage V' in rt: emission = 'Stage V'
                elif 'Stage IV' in rt: emission = 'Stage IV'
                elif 'Stage III' in rt: emission = 'Stage III'
                elif 'Stage II' in rt: emission = 'Stage II'
                elif 'Tier 4' in rt: emission = 'Tier 4F'
            
            if fob is None and dap is None:
                for i, row in enumerate(rows):
                    p = find_price(row, 1)
                    if p and 1 <= i <= 5:
                        if fob is None:
                            fob = p
                            desc_en = safe_str(row[0]) if row[0] else ''
                        elif dap is None and p != fob:
                            dap = p
                            break
            
            if not desc_en or len(desc_en) < 20:
                for row in rows:
                    r0 = safe_str(row[0])
                    if len(r0) > 100 and ('/' in r0 or 'EXCAVATOR' in r0.upper()):
                        desc_en = r0
                        desc_cn = safe_str(row[1]) if len(row) > 1 else ''
                        break
            
            if fob or dap:
                eng, hp = '', 0
                text = (desc_en + ' ' + desc_cn).lower()
                for b in ['cummins','yuchai','yunnei','yanmar','weichai','deutz','isuzu']:
                    if b in text: eng = b.capitalize(); break
                hm = re.search(r'(\d+)\s*hp', text)
                if hm: hp = int(hm.group(1))
                km = re.search(r'(\d+)\s*kw', text)
                if km and not hp: hp = int(int(km.group(1)) * 1.36)
                
                bm = re.search(r'(\d+\.?\d*)\s*[mMмМ][3³]', text)
                if not bm: bm = re.search(r'(\d+\.?\d*)\s*方', text)
                bucket = float(bm.group(1)) if bm else 0
                
                products.append({
                    'model': sn.strip(),
                    'category': determine_category(filename),
                    'desc_ru': desc_en[:300],
                    'desc_cn': desc_cn[:300],
                    'fob_price_cny': fob,
                    'dap_price_cny': dap if dap else fob,
                    'engine': eng,
                    'hp': hp,
                    'emission': emission,
                    'bucket_m3': bucket,
                    'source_file': filename,
                    'source_sheet': sn,
                })
        except Exception as e:
            print(f'  Sheet {sn}: {e}')
    wb.close()
    return products

def load_scrap_tax(filepath):
    tax = {}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = None
    
    # Check for RL consolidated sheet format (LGRU-RUSSIA-RL)
    rl_sheet = None
    for sn in wb.sheetnames:
        snu = sn.upper().replace(' ', '').replace('-', '')
        if 'LGRURUSSIARL' in snu:
            rl_sheet = sn
            break
    
    if rl_sheet:
        ws = wb[rl_sheet]
        rows_list = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        # Find header row with FOB and DAP
        hdr_row = -1
        for i, row in enumerate(rows_list):
            rt = ' '.join([safe_str(c) for c in row])
            if 'FOB' in rt.upper() and 'DAP' in rt.upper():
                hdr_row = i
                break
        if hdr_row >= 0:
            hdr = [safe_str(c) for c in rows_list[hdr_row]]
            model_col = fob_col = dap_col = -1
            for j, h in enumerate(hdr):
                hu = h.upper()
                if 'MODEL' in hu or '型号' in h:
                    model_col = j
                if 'FOB' in hu:
                    fob_col = j
                if 'DAP' in hu:
                    dap_col = j
            
            for i in range(hdr_row + 1, len(rows_list)):
                row = rows_list[i]
                model = safe_str(row[model_col]) if model_col >= 0 and model_col < len(row) else ''
                if not model or len(model) < 3:
                    continue
                fob_rl = None
                dap_rl = None
                try:
                    if fob_col >= 0 and fob_col < len(row):
                        fob_rl = float(row[fob_col]) if row[fob_col] else None
                except:
                    pass
                try:
                    if dap_col >= 0 and dap_col < len(row):
                        dap_rl = float(row[dap_col]) if row[dap_col] else None
                except:
                    pass
                
                if fob_rl and fob_rl > 1000:
                    desc_en_rl = safe_str(row[4]) if len(row) > 4 else ''
                    desc_cn_rl = safe_str(row[3]) if len(row) > 3 else ''
                    eng, hp = parse_eng_hp(desc_en_rl + ' ' + desc_cn_rl)
                    products.append({
                        'model': model,
                        'category': determine_category(filename, sn),
                        'desc_ru': desc_en_rl[:300],
                        'desc_cn': desc_cn_rl[:300],
                        'fob_price_cny': fob_rl,
                        'dap_price_cny': dap_rl if dap_rl else fob_rl,
                        'engine': eng,
                        'hp': hp,
                        'emission': '',
                        'bucket_m3': 0,
                        'tonnage': 0,
                        'source_file': filename,
                        'source_sheet': rl_sheet,
                    })
        wb.close()
        return products

    for sn in wb.sheetnames:
            if 'Latest' in sn: sheet = sn; break
        if not sheet:
        
    # Check for RL consolidated sheet format (LGRU-RUSSIA-RL)
    rl_sheet = None
    for sn in wb.sheetnames:
        snu = sn.upper().replace(' ', '').replace('-', '')
        if 'LGRURUSSIARL' in snu:
            rl_sheet = sn
            break
    
    if rl_sheet:
        ws = wb[rl_sheet]
        rows_list = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        # Find header row with FOB and DAP
        hdr_row = -1
        for i, row in enumerate(rows_list):
            rt = ' '.join([safe_str(c) for c in row])
            if 'FOB' in rt.upper() and 'DAP' in rt.upper():
                hdr_row = i
                break
        if hdr_row >= 0:
            hdr = [safe_str(c) for c in rows_list[hdr_row]]
            model_col = fob_col = dap_col = -1
            for j, h in enumerate(hdr):
                hu = h.upper()
                if 'MODEL' in hu or '型号' in h:
                    model_col = j
                if 'FOB' in hu:
                    fob_col = j
                if 'DAP' in hu:
                    dap_col = j
            
            for i in range(hdr_row + 1, len(rows_list)):
                row = rows_list[i]
                model = safe_str(row[model_col]) if model_col >= 0 and model_col < len(row) else ''
                if not model or len(model) < 3:
                    continue
                fob_rl = None
                dap_rl = None
                try:
                    if fob_col >= 0 and fob_col < len(row):
                        fob_rl = float(row[fob_col]) if row[fob_col] else None
                except:
                    pass
                try:
                    if dap_col >= 0 and dap_col < len(row):
                        dap_rl = float(row[dap_col]) if row[dap_col] else None
                except:
                    pass
                
                if fob_rl and fob_rl > 1000:
                    desc_en_rl = safe_str(row[4]) if len(row) > 4 else ''
                    desc_cn_rl = safe_str(row[3]) if len(row) > 3 else ''
                    eng, hp = parse_eng_hp(desc_en_rl + ' ' + desc_cn_rl)
                    products.append({
                        'model': model,
                        'category': determine_category(filename, sn),
                        'desc_ru': desc_en_rl[:300],
                        'desc_cn': desc_cn_rl[:300],
                        'fob_price_cny': fob_rl,
                        'dap_price_cny': dap_rl if dap_rl else fob_rl,
                        'engine': eng,
                        'hp': hp,
                        'emission': '',
                        'bucket_m3': 0,
                        'tonnage': 0,
                        'source_file': filename,
                        'source_sheet': rl_sheet,
                    })
        wb.close()
        return products

    for sn in wb.sheetnames:
                ws = wb[sn]
                r0 = [safe_str(c) for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
                if any('26' in c for c in r0): sheet = sn; break
        if sheet:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                vals = [safe_str(c) for c in row]
                if len(vals) >= 9 and vals[1]:
                    m = vals[1].strip()
                    try:
                        tax_rub = float(vals[8]) if vals[8] else 0
                    except: tax_rub = 0
                    try: hp_v = float(vals[4]) if len(vals)>4 and vals[4] else 0
                    except: hp_v = 0
                    tax[m] = {'scrap_tax_2026_rub': tax_rub, 'hp': int(hp_v)}
        wb.close()
    except Exception as e:
        print(f'Scrap tax error: {e}')
    return tax

print('Building LiuGong product database...')
tax_map = load_scrap_tax(os.path.join(PRODUCTS_DIR, '2026年报废税 2026Y scrap tax.xlsx'))
print(f'Loaded scrap tax for {len(tax_map)} models')

all_prods = []
for fname in sorted(os.listdir(PRODUCTS_DIR)):
    if not fname.endswith(('.xlsx','.xls')): continue
    if '报废税' in fname: continue
    fpath = os.path.join(PRODUCTS_DIR, fname)
    print(f'Processing: {fname[:50]}...')
    prods = extract_products(fpath, fname)
    for p in prods:
        m = p['model']
        if m in tax_map:
            p['scrap_tax_rub'] = tax_map[m]['scrap_tax_2026_rub']
            if not p['hp']: p['hp'] = tax_map[m]['hp']
        else:
            p['scrap_tax_rub'] = 0
    all_prods.extend(prods)
    print(f'  -> {len(prods)} products')

seen = {}
unique = []
for p in all_prods:
    m = p['model']
    if m in seen:
        if p['dap_price_cny'] and not seen[m]['dap_price_cny']:
            seen[m] = p
        continue
    seen[m] = p
unique = sorted(seen.values(), key=lambda p: (p['category'], p['model']))

cats = {}
for p in unique:
    c = p['category']
    cats[c] = cats.get(c, 0) + 1

output = {'meta': {'version':'20260611-full','total':len(unique),'categories':cats}, 'products': unique}
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f'Done! {len(unique)} unique products')
for c,n in sorted(cats.items()):
    print(f'  {c}: {n}')
print(f'Output: {OUTPUT_PATH}')
