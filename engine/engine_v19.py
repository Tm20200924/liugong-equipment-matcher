# -*- coding: utf-8 -*-
"""LiuGong Equipment Matching Engine v19 - Clean & Focused"""
import os, json, re, sys
from datetime import datetime
import openpyxl
try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INQUIRIES_DIR = os.path.join(BASE_DIR, "inquiries")
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")
CONFIG_PATH = os.path.join(BASE_DIR, "engine", "config.json")
DB_PATH = os.path.join(BASE_DIR, "engine", "product_db_full.json")
COMP_DB_PATH = os.path.join(BASE_DIR, "engine", "competitor_db.json")
os.makedirs(OUTPUTS_DIR, exist_ok=True)

# ===== CONFIG & DATA LOADING =====

def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8-sig") as f:
            return json.load(f)
    return {"exchange_rate_rub_cny": 11.5, "customs_duty_rate_pct": 5.0,
            "vat_rate_pct": 22.0}

def load_db():
    if os.path.exists(DB_PATH):
        with open(DB_PATH, "r", encoding="utf-8") as f:
            products = json.load(f).get("products", [])
        return [p for p in products if p.get("model","").strip() and not str(p.get("model","")).startswith("Sheet") and "Price" not in str(p.get("model","")) and "??" not in str(p.get("model","")) and "??" not in str(p.get("model","")) and "???" not in str(p.get("model","")) and p.get("dap_price_cny") and p.get("dap_price_cny", 0) > 0]
    return []

def load_competitor_db():
    if os.path.exists(COMP_DB_PATH):
        with open(COMP_DB_PATH, "r", encoding="utf-8") as f:
            return json.load(f).get("known", {})
    return {}

# ===== ONLINE SEARCH (optional, falls back to local) =====

ONLINE_CACHE = {}  # Session cache to avoid repeat searches
ONLINE_TIMEOUT = 8  # seconds

def search_online_specs(query, is_competitor=True):
    """Search online for equipment specifications.
    Returns dict with {tonnage, hp, bucket_m3, source} or None."""
    if not HAS_REQUESTS:
        return None
    
    cache_key = query.lower().strip()
    if cache_key in ONLINE_CACHE:
        return ONLINE_CACHE[cache_key]
    
    try:
        # Use DuckDuckGo HTML (no API key, more reliable than Google scraping)
        url = "https://html.duckduckgo.com/html/"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        data = {"q": query}
        resp = requests.post(url, data=data, headers=headers, timeout=ONLINE_TIMEOUT)
        html = resp.text
        
        result = {"source": "web"}
        
        # Extract snippets from search results
        snippets = re.findall(r'class="result__snippet"[^>]*>(.*?)</a>', html, re.DOTALL)
        if not snippets:
            snippets = re.findall(r'<[^>]*class="[^"]*snippet[^"]*"[^>]*>(.*?)</', html, re.DOTALL)
        all_text = " ".join(snippets) if snippets else html
        
        # Clean HTML tags
        all_text = re.sub(r'<[^>]+>', ' ', all_text)
        all_text = re.sub(r'\s+', ' ', all_text)
        
        # Extract tonnage (operating weight)
        ton = 0
        for pat in [
            r'(\d+\.?\d*)\s*(?:metric\s*)?(?:ton|t)[s\b]',
            r'(\d+\.?\d*)\s*[tT]\b',
            r'(?:operating|working|machine)\s*weight[:\s]*(\d+[\.,]?\d*)\s*(?:kg|KG|ton)',
        ]:
            tm = re.search(pat, all_text, re.IGNORECASE)
            if tm:
                v = float(tm.group(1).replace(",", "."))
                if v > 100:
                    v = v / 1000.0
                if 0.5 < v < 200:
                    ton = round(v, 1)
                    break
        
        # Extract horsepower
        hp = 0
        for pat in [
            r'(\d+)\s*(?:HP|hp|Hp)\b',
            r'(\d+)\s*(?:kW|kw|Kw)\b',
            r'(?:engine|net)\s*power[:\s]*(\d+)\s*(?:HP|hp|kW)',
        ]:
            hm = re.search(pat, all_text, re.IGNORECASE)
            if hm:
                v = int(hm.group(1))
                if "kW" in pat.lower():
                    v = int(v * 1.341)
                if 5 < v < 2000:
                    hp = v
                    break
        
        # Extract bucket capacity
        bucket = 0
        for pat in [
            r'(\d+\.?\d*)\s*m[3³]',
            r'bucket\s*(?:capacity|size)[:\s]*(\d+\.?\d*)\s*m',
        ]:
            bm = re.search(pat, all_text, re.IGNORECASE)
            if bm:
                v = float(bm.group(1))
                if 0.01 < v < 20:
                    bucket = round(v, 2)
                    break
        
        if ton > 0 or hp > 0:
            result["tonnage"] = ton
            result["hp"] = hp
            result["bucket_m3"] = bucket
            ONLINE_CACHE[cache_key] = result
            print(f"  [Web] {query[:50]:50s} -> ton={ton} hp={hp} bucket={bucket}")
            return result
        
    except Exception as e:
        print(f"  [Web] Search failed ({type(e).__name__}), using local DB")
    
    ONLINE_CACHE[cache_key] = None
    return None

def verify_specs_online(model, category, local_ton=0, local_hp=0, local_bucket=0):
    """Cross-validate product specs: online vs local DB.
    Returns {level, online_ton, online_hp, online_bucket, local_ton, local_hp, local_bucket, status}."""
    result = {
        "level": "none", "online_ton": 0, "online_hp": 0, "online_bucket": 0,
        "local_ton": local_ton, "local_hp": local_hp, "local_bucket": local_bucket,
        "status": "no_data"
    }
    online = search_online_specs(f"LiuGong {model} specifications weight horsepower")
    if online:
        ot = online.get("tonnage", 0) or 0
        oh = online.get("hp", 0) or 0
        ob = online.get("bucket_m3", 0) or 0
        result["online_ton"] = ot
        result["online_hp"] = oh
        result["online_bucket"] = ob
        has_local = local_ton > 0 or local_hp > 0
        has_online = ot > 0 or oh > 0
        if has_local and has_online:
            conflicts = []
            if ot > 0 and local_ton > 0:
                diff_pct = abs(ot - local_ton) / max(local_ton, 0.1)
                if diff_pct < 0.05:
                    pass  # match
                elif diff_pct < 0.10:
                    conflicts.append(f"ton_diff_{diff_pct:.0%}")
                else:
                    conflicts.append(f"ton_conflict_{diff_pct:.0%}")
            if oh > 0 and local_hp > 0:
                diff_pct = abs(oh - local_hp) / max(local_hp, 0.1)
                if diff_pct < 0.05:
                    pass
                elif diff_pct < 0.10:
                    conflicts.append(f"hp_diff_{diff_pct:.0%}")
                else:
                    conflicts.append(f"hp_conflict_{diff_pct:.0%}")
            if conflicts:
                if any("conflict" in c for c in conflicts):
                    result["level"] = "conflict"
                    result["status"] = "conflict: " + ", ".join(conflicts)
                else:
                    result["level"] = "close"
                    result["status"] = "close: " + ", ".join(conflicts)
            else:
                result["level"] = "dual"
                result["status"] = "dual_verified"
        elif has_online:
            result["level"] = "single"
            result["status"] = "online_only"
        else:
            result["level"] = "single"
            result["status"] = "local_only"
    else:
        if local_ton > 0 or local_hp > 0:
            result["level"] = "single"
            result["status"] = "local_only"
    return result

def search_competitor_online(brand, model):
    """Search online for competitor specs. Returns {tonnage, hp, bucket} or None."""
    queries = [
        f"{brand} {model} specifications operating weight horsepower bucket",
        f"{brand} {model} excavator specs ton weight",
        f"{brand} {model} bulldozer loader specs weight",
    ]
    for q in queries:
        result = search_online_specs(q)
        if result:
            return result
    return None

# ===== CATEGORY SYSTEM =====

CAT_CN = {
    "EX": "挖掘机", "WL": "装载机", "RL": "压路机", "BD": "推土机",
    "MG": "平地机", "SSL": "滑移装载机", "BHL": "两头忙", "MT": "矿卡",
    "AWP": "高空作业平台", "TRUCK": "汽车吊", "TELEHANDLER": "叉装车", "ELECTRIC": "电动产品",
}

# Keywords for matching inquiry text to LiuGong equipment categories
CATEGORY_KEYWORDS = {
    "挖掘机": ["挖掘", "挖机", "excavator", "экскаватор", "轮挖"],
    "装载机": ["装载", "loader", "погрузчик", "铲车", "фронтальный"],
    "推土机": ["推土", "bulldozer", "dozer", "бульдозер"],
    "压路机": ["压路", "roller", "каток", "виброкаток", "振动碾", "单钢轮", "双钢轮", "轮胎压路"],
    "平地机": ["平地", "grader", "грейдер"],
    "两头忙": ["两头忙", "backhoe", "экскаватор-погрузчик"],
    "矿卡": ["矿卡", "mining truck", "самосвал", "карьерный", "自卸车", "dump truck"],
    "汽车吊": ["汽车吊", "truck crane", "автокран", "吊车"],
    "滑移装载机": ["滑移", "skid steer", "мини-погрузчик"],
    "叉装车": ["叉装", "telehandler", "телескопический"],
    "电动产品": ["电动装载机", "电动挖掘机", "electric loader", "electric excavator"],
    "高空作业平台": ["高空作业", "AWP", "подъемник", "升降"],
}

# Equipment LiuGong does NOT produce (for display purposes)
NON_LIUGONG = ["洒水车", "摊铺机", "焊机", "搅拌机", "叉车", "水泵", "发电机"]

# ===== COMPETITOR DETECTION =====

# Model prefix -> (brand, category)
COMPETITOR_PREFIX = {
    "XE": ("XCMG", "挖掘机"), "SY": ("SANY", "挖掘机"), "ZE": ("Zoomlion", "挖掘机"),
    "PC": ("Komatsu", "挖掘机"), "ZX": ("Hitachi", "挖掘机"), "EC": ("Volvo", "挖掘机"),
    "DX": ("Doosan", "挖掘机"), "CAT": ("Caterpillar", "挖掘机"),
    "SK": ("Kobelco", "挖掘机"), "LG": ("SDLG", "挖掘机"),
    "LW": ("XCMG", "装载机"), "ZL": ("XCMG", "装载机"), "SW": ("SANY", "装载机"),
    "TY": ("XCMG", "推土机"), "SD": ("SHANTUI", "推土机"),
    "D6": ("Caterpillar", "推土机"), "D7": ("Caterpillar", "推土机"),
    "D8": ("Caterpillar", "推土机"), "D9": ("Caterpillar", "推土机"),
    "D10": ("Caterpillar", "推土机"), "D11": ("Caterpillar", "推土机"),
    "XS": ("XCMG", "压路机"), "SSR": ("SANY", "压路机"), "SR": ("SANY", "压路机"),
    "QY": ("XCMG", "汽车吊"), "STC": ("SANY", "汽车吊"),
}

# ===== TEXT EXTRACTION =====

def classify_inquiry(name):
    nl = name.lower()
    for cat, kws in CATEGORY_KEYWORDS.items():
        for kw in kws:
            if kw.lower() in nl:
                return cat
    return None

def is_non_liugong(name):
    """Whitelist: only match if inquiry has LiuGong category keywords or competitor model."""
    if classify_inquiry(name):
        return False
    comp, _ = detect_competitor(name)
    if comp:
        return False
    return True

def extract_ton(text):
    m = re.search(r"(\d+\.?\d*)\s*[tT吨tonTONТт]", text)
    if m: return float(m.group(1))
    m = re.search(r"(\d+)\s*тонн", text, re.IGNORECASE)
    if m: return float(m.group(1))
    return 0

def extract_bucket(text):
    m = re.search(r"(\d+\.?\d*)\s*[mM]\s*[3³]", text)
    if not m: m = re.search(r"(\d+\.?\d*)\s*方", text)
    if not m: m = re.search(r"(\d+\.?\d*)\s*куб", text, re.IGNORECASE)
    return float(m.group(1)) if m else 0

def extract_hp(text):
    m = re.search(r"(\d+)\s*(?:hp|马力|л\.с)", text, re.IGNORECASE)
    if m: return int(m.group(1))
    m = re.search(r"(\d+)\s*(?:kw|кВт)", text, re.IGNORECASE)
    if m: return int(int(m.group(1)) * 1.36)
    return 0

def detect_competitor(text):
    """Detect competitor brand from text. Returns (brand, model_hint) or (None, None)."""
    # Clean and search for prefix patterns
    tu = re.sub(r'[()\-\s/:：;；]', '', text.upper())
    for prefix, (brand, cat) in sorted(COMPETITOR_PREFIX.items(), key=lambda x: -len(x[0])):
        idx = tu.find(prefix)
        if idx >= 0:
            nidx = idx + len(prefix)
            if nidx < len(tu) and tu[nidx].isdigit():
                # Extract the model number
                m = re.search(r'([A-Z]{2,4}\s*\d{2,4})', text.upper())
                model = m.group(1).replace(" ", "") if m else ""
                return brand, model
    return None, ""

def extract_model_ton_from_text(text, comp_db=None):
    """Extract tonnage from model: checks competitor DB first, then pattern.
    TY230=23t, SD26=26t, CAT320=20t (from DB), PC200=20t."""
    if comp_db is None:
        comp_db = {}
    # 1. Check competitor DB for exact match
    tu = re.sub(r'[()\-\.\/\s]', '', text.upper())
    for key, (brand, ton, bucket) in comp_db.items():
        kc = re.sub(r'[()\-\.\/\s]', '', key.upper())
        if kc == tu or kc in tu or tu in kc:
            if ton > 0:
                return ton
    # 2. Pattern-based fallback
    ms = re.findall(r'([A-Za-z]{2,4})\s*(\d{2,4})', text.upper().replace("-", ""))
    for prefix, nums in ms:
        v = int(nums)
        if 100 <= v < 1000:
            return v / 10.0
        elif 3 < v < 100:
            return float(v)
    return 0

# ===== MODEL TONNAGE =====

def model_tonnage(model):
    """Extract tonnage from LiuGong model number."""
    m = str(model).upper().replace(" ", "").replace("CLG", "")
    md = re.sub(r"[A-Z\-]+$", "", m)
    
    # Truck crane: LTC250=25t
    ltc = re.match(r"LTC(\d{2,3})", md)
    if ltc:
        v = int(ltc.group(1))
        return v / 10.0 if v >= 100 else float(v)
    
    # Dozer LD: LD20=20t
    bd = re.match(r"LD(\d{2})", md)
    if bd: return float(bd.group(1))
    
    # Dozer B: B260=26t, B160=16t
    bd2 = re.match(r"B(\d{2,3})", md)
    if bd2:
        v = int(bd2.group(1))
        return v / 10.0 if v > 50 else float(v)
    
    # DW series
    dw = re.match(r"DW(\d{2,3})", md)
    if dw: return float(dw.group(1))
    
    # Mini excavators: 9018=1.8t, 908EN=8t
    m90 = re.match(r"^90(\d{2})", md)
    if m90:
        v = int(m90.group(1))
        return v / 10.0 if v <= 60 else float(v)
    
    # 9-series excavators: 920E=20t, 950E=50t
    m9 = re.match(r"^9(\d{2,3})", md)
    if m9:
        v = int(m9.group(1))
        return float(v) if v < 50 else float(v)
    
    # Wheel loaders: 835H=3.5t, 870H=7.0t
    l8 = re.match(r"^8(\d)(\d)", md)
    if l8:
        return float(l8.group(1)) + float(l8.group(2)) / 10.0
    
    # Rollers: 6xxx
    # 60xx=tandem(small), 61xx=single drum, 62xx=single drum, 65xx=pneumatic, 66xx=double drum
    r6 = re.match(r"^6(\d)(\d)(\d)", md)
    if r6:
        prefix = int(r6.group(1))
        last2 = int(r6.group(2) + r6.group(3))
        if prefix in [1, 2, 5, 6] and last2 <= 50:
            return float(last2)
        if prefix == 0:
            return last2 / 10.0
    
    # Graders: 4150=15t, 4215=21.5t
    g4 = re.match(r"^4(\d{2,3})", md)
    if g4:
        v = int(g4.group(1))
        return v / 10.0 if v < 200 else float(v)
    
    return 0

# ===== BUCKET TO TONNAGE (excavator/loader only) =====

BUCKET_EX = {0.04:1.8, 0.06:2.7, 0.08:3.5, 0.1:5.5, 0.2:6, 0.3:8, 0.4:9,
    0.5:13, 0.6:15, 0.7:18, 0.8:20, 0.9:22, 1.0:23, 1.2:26,
    1.4:30, 1.6:33, 1.8:36, 2.0:39, 2.5:50, 3.0:65, 3.5:75, 4.0:95}

BUCKET_WL = {0.8:1.6, 1.0:1.8, 1.2:2.0, 1.5:2.0, 1.6:2.0, 1.8:3.0,
    2.0:3.0, 2.5:3.5, 3.0:5.0, 3.5:5.5, 4.0:6.5, 5.0:8.0}

def bucket_to_ton(bucket, cat):
    if cat == "挖掘机":
        for b, t in sorted(BUCKET_EX.items()):
            if bucket <= b: return t
        return bucket * 20
    if cat == "装载机":
        for b, t in sorted(BUCKET_WL.items()):
            if bucket <= b: return t
        return bucket * 1.5
    return 0

# ===== ROLLER SUB-CATEGORY =====

def roller_sub(model):
    m = str(model).upper().replace("CLG", "")
    md = re.sub(r"[A-Z\-]+$", "", m)
    rm = re.match(r"^6(\d)(\d)(\d)", md)
    if rm:
        p = int(rm.group(1))
        if p == 0: return "小型压路机"
        if p in [1, 2]: return "单钢轮压路机"
        if p == 5: return "轮胎压路机"
        if p == 6: return "双钢轮压路机"
    return None

def roller_hint(text):
    txt = text.lower()
    if any(k in txt for k in ["小型","双钢轮","双光轮","手扶","tandem"]): return "小型压路机"
    if any(k in txt for k in ["轮胎压路","轮胎式","胶轮","pneumatic"]): return "轮胎压路机"
    if any(k in txt for k in ["单钢轮","振动","单驱"]): return "单钢轮压路机"
    if any(k in txt for k in ["双驱","双轮"]): return "双钢轮压路机"
    return None

# ===== CORE MATCHING =====

def match_products(name, db, config):
    """Match inquiry to LiuGong products. Returns (candidates, ton, bucket, verification)."""
    
    # Reject non-LiuGong equipment
    if is_non_liugong(name):
        return [], 0, 0, None
    
    # Detect competitor
    comp_db = load_competitor_db()
    comp_brand, comp_model = detect_competitor(name)
    if comp_brand:
        # Competitor: extract tonnage from competitor DB first, then model, then text
        # Step 1: Try online search first (most accurate)
        comp_ton = 0
        online_specs = None
        if comp_brand and comp_model:
            online_specs = search_competitor_online(comp_brand, comp_model)
            if online_specs and online_specs.get("tonnage", 0) > 0:
                comp_ton = online_specs["tonnage"]
        # Step 2: Cross-validate with local DB
        local_ton = extract_ton(name) or extract_model_ton_from_text(name, comp_db) or 0
        if local_ton > 0 and comp_ton > 0:
            # If both available, use average (cross-validation)
            if abs(local_ton - comp_ton) < 5:
                comp_ton = (comp_ton + local_ton) / 2.0
            elif abs(local_ton - comp_ton) < 10:
                comp_ton = local_ton  # Trust local more when close
        elif local_ton > 0 and not comp_ton:
            comp_ton = local_ton
        comp_cat = None
        tu = re.sub(r'[()\-\s/:：;；]', '', comp_model.upper())
        for prefix, (brand, cat) in COMPETITOR_PREFIX.items():
            if tu.startswith(prefix):
                comp_cat = cat; break
        
        candidates = []
        for p in db:
            pcat = CAT_CN.get(p["category"], p["category"])
            if comp_cat and pcat != comp_cat:
                continue
            pton = model_tonnage(p["model"]) or p.get("tonnage", 0)
            score = 0
            reasons = [f"竞品对标: {comp_brand} {comp_model}"]
            
            if comp_ton > 0 and pton > 0:
                diff_pct = abs(comp_ton - pton) / max(comp_ton, 0.1)
                if comp_ton < 5:
                    if diff_pct <= 0.10: score += 50
                    elif diff_pct <= 0.20: score += 35
                    elif diff_pct <= 0.35: score += 15
                    else: score -= 15
                elif comp_ton < 20:
                    if diff_pct <= 0.10: score += 50
                    elif diff_pct <= 0.25: score += 35
                    elif diff_pct <= 0.45: score += 15
                    else: score -= 10
                elif comp_ton < 50:
                    if diff_pct <= 0.15: score += 50
                    elif diff_pct <= 0.30: score += 35
                    elif diff_pct <= 0.50: score += 15
                    else: score -= 5
                else:
                    if diff_pct <= 0.20: score += 50
                    elif diff_pct <= 0.40: score += 35
                    elif diff_pct <= 0.60: score += 15
                    else: score -= 5
                if score > 0:
                    reasons.append(f"吨位({pton}t,竞品{comp_ton}t,差{diff_pct:.0%})")
            
            if p.get("dap_price_cny"): score += 10
            candidates.append({"product": p, "score": score, "reasons": reasons, "pton": pton})
        
        candidates.sort(key=lambda x: (x["score"], -abs(comp_ton - x["pton"]) if comp_ton > 0 and x["pton"] > 0 else 0), reverse=True)
        if candidates and candidates[0]["score"] > 5:
            verification = {
            "query": name, "classified_cat": comp_cat, "extracted_ton": comp_ton,
            "competitor": comp_brand, "competitor_model": comp_model,
            "online_search_used": True, "top_verification": None
        }
        if candidates:
            top = candidates[0]
            p = top["product"]
            pton = model_tonnage(p["model"]) or p.get("tonnage", 0)
            verification["top_verification"] = verify_specs_online(
                p["model"], comp_cat or "", local_ton=pton
            )
        return candidates, comp_ton, extract_bucket(name), verification
        # If competitor match fails, fall through to normal matching
    
    # Normal LiuGong matching
    cat = classify_inquiry(name)
    ton = extract_ton(name)
    bucket = extract_bucket(name)
    hp_req = extract_hp(name)
    
    # Try model-based tonnage if no explicit ton
    if not ton:
        ton = extract_model_ton_from_text(name, comp_db)
    
    # Bucket-to-ton only for excavator/loader
    if not ton and bucket and cat in ("挖掘机", "装载机"):
        ton = bucket_to_ton(bucket, cat)
    
    candidates = []
    for p in db:
        pcat = CAT_CN.get(p["category"], p["category"])
        if cat and pcat != cat:
            continue
        
        pton = model_tonnage(p["model"]) or p.get("tonnage", 0)
        score = 0
        reasons = []
        
        # Tonnage scoring
        if ton > 0 and pton > 0:
            diff_pct = abs(ton - pton) / max(ton, 0.1)
            if ton < 5:
                if diff_pct <= 0.10: score += 50; reasons.append(f"吨位精准匹配({pton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.20: score += 35; reasons.append(f"吨位接近({pton}t,需求{ton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.35: score += 15; reasons.append(f"吨位可参考({pton}t,差{diff_pct:.0%})")
                else: score -= 15; reasons.append(f"吨位不匹配({pton}t,差{diff_pct:.0%})")
            elif ton < 20:
                if diff_pct <= 0.10: score += 50; reasons.append(f"吨位精准匹配({pton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.25: score += 35; reasons.append(f"吨位接近({pton}t,需求{ton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.45: score += 15; reasons.append(f"吨位可参考({pton}t,差{diff_pct:.0%})")
                else: score -= 10; reasons.append(f"吨位不匹配({pton}t,差{diff_pct:.0%})")
            elif ton < 50:
                if diff_pct <= 0.15: score += 50; reasons.append(f"吨位精准匹配({pton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.30: score += 35; reasons.append(f"吨位接近({pton}t,需求{ton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.50: score += 15; reasons.append(f"吨位可参考({pton}t,差{diff_pct:.0%})")
                else: score -= 5
            else:
                if diff_pct <= 0.20: score += 50; reasons.append(f"吨位精准匹配({pton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.40: score += 35; reasons.append(f"吨位接近({pton}t,需求{ton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.60: score += 15; reasons.append(f"吨位可参考({pton}t,差{diff_pct:.0%})")
                else: score -= 5
        
        # Bucket scoring (excavator/loader only)
        pbucket = p.get("bucket_m3", 0)
        if bucket > 0 and pbucket > 0 and cat in ("挖掘机", "装载机"):
            bdiff = abs(bucket - pbucket)
            if bdiff < 0.1: score += 30; reasons.append(f"斗容匹配({pbucket}m3)")
            elif bdiff < 0.5: score += 15; reasons.append(f"斗容接近({pbucket}m3)")
        
        # Horsepower scoring
        ph = p.get("hp", 0)
        if hp_req > 0 and ph > 0:
            hdiff = abs(hp_req - ph) / max(hp_req, 1)
            if hdiff < 0.1: score += 20; reasons.append(f"马力匹配({ph}hp)")
            elif hdiff < 0.3: score += 10; reasons.append(f"马力接近({ph}hp)")
        
        if p.get("dap_price_cny"): score += 10
        if cat and pcat == cat: score += 5
        # Exact model mention boost
        model_clean = re.sub(r"[^A-Za-z0-9]", "", str(p.get("model", "")))
        name_clean = re.sub(r"[^A-Za-z0-9]", "", name.upper())
        if model_clean and len(model_clean) >= 3 and name_clean and len(name_clean) >= 3 and not name_clean.isdigit() and (model_clean in name_clean or name_clean in model_clean):
            score += 40
            reasons.append(f"型号直接提及({p["model"]})")
        
        candidates.append({"product": p, "score": score, "reasons": reasons, "pton": pton})
    
    candidates.sort(key=lambda x: x["score"], reverse=True)
    
    # Roller sub-category boost
    if cat == "压路机" and candidates:
        hint = roller_hint(name)
        if hint:
            for c in candidates:
                sub = roller_sub(c["product"]["model"])
                if sub == hint:
                    c["score"] += 20
                    c["reasons"].append(f"子类匹配({sub})")
            candidates.sort(key=lambda x: x["score"], reverse=True)
    
    # Build verification chain
    verification = {
        "query": name,
        "classified_cat": cat,
        "extracted_ton": ton,
        "extracted_hp": hp_req,
        "extracted_bucket": bucket,
        "competitor": comp_brand,
        "competitor_model": comp_model,
        "online_search_used": bool(comp_brand and comp_model),
        "top_verification": None
    }
    if candidates:
        top = candidates[0]
        p = top["product"]
        pton = model_tonnage(p["model"]) or p.get("tonnage", 0)
        ph = p.get("hp", 0) or 0
        pb = p.get("bucket_m3", 0) or 0
        verification["top_verification"] = verify_specs_online(
            p["model"], cat or "",
            local_ton=pton, local_hp=ph, local_bucket=pb
        )
    return candidates, ton, bucket, verification

# ===== DAP CALCULATION =====

def calc_dap(dap_cny, scrap_rub, config):
    rate = config.get("exchange_rate_rub_cny", 11.5)
    duty_pct = config.get("customs_duty_rate_pct", 5.0)
    proc_pct = config.get("customs_processing_rate_pct", 0.3)
    vat_pct = config.get("vat_rate_pct", 22.0)
    warehouse = config.get("customs_warehouse_fee_rub", 20000)
    customs_fee = config.get("customs_fee_rub", 6000)
    agent_fee = config.get("customs_agent_fee_rub", 30000)
    
    rub_fixed = scrap_rub + warehouse + customs_fee + agent_fee
    cny_duty = dap_cny * duty_pct / 100.0
    cny_proc = dap_cny * proc_pct / 100.0
    rub_fixed_to_cny = rub_fixed / rate
    vat_base = dap_cny + cny_duty + cny_proc + rub_fixed_to_cny
    vat_cny = vat_base * vat_pct / 100.0
    total_rub = (vat_base + vat_cny) * rate
    
    return {
        "dap_cny": round(dap_cny, 2), "rate": rate,
        "duty_pct": duty_pct, "duty_rub": round(cny_duty * rate, 2),
        "proc_pct": proc_pct, "proc_rub": round(cny_proc * rate, 2),
        "scrap_rub": round(scrap_rub, 2), "warehouse_rub": round(warehouse, 2),
        "customs_fee_rub": round(customs_fee, 2), "agent_fee_rub": round(agent_fee, 2),
        "fixed_rub": round(rub_fixed, 2),
        "vat_pct": vat_pct, "vat_rub": round(vat_cny * rate, 2),
        "total_rub": round(total_rub, 2),
    }

# ===== EXCEL PARSING =====

def parse_inquiry_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=True))
    inquiries = []
    
    for i, row in enumerate(rows):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if not any(vals): continue
        
        all_text = " ".join(vals).lower()
        # Skip header rows
        if any(h in all_text for h in ["序号", "设备名称", "名称", "№"]): continue
        # Skip category-only rows (no sequence number)
        has_seq = vals[0] and vals[0].replace(".", "").replace("-", "").isdigit()
        if not has_seq and len([v for v in vals if v]) <= 2: continue
        
        name = ""
        qty = 1
        for v in vals:
            if not v: continue
            if not name and len(v) > 2:
                is_equip = any(kw.lower() in v.lower() for kws in CATEGORY_KEYWORDS.values() for kw in kws)
                is_comp = detect_competitor(v)[0]
                is_other = not v.replace(".","").replace(",","").replace("-","").isdigit()
                if is_equip or is_comp or (is_other and len(v) > 3):
                    name = v
            try:
                q = int(v)
                if 1 <= q <= 1000 and v == str(q):
                    qty = q
            except: pass
        
        if name:
            inquiries.append({"seq": str(i+1), "name": name, "qty": qty})
    
    wb.close()
    return inquiries

# ===== REPORT GENERATION =====

def generate_report(inquiries, db, config, fname):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "匹配结果"
    
    hdr_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    headers = [
        "序号", "客户询盘需求", "数量", "设备类别",
        "推荐型号", "产品俄语全称", "推荐理由", "匹配度",
        "DAP满洲里(CNY)", "汇率(CNY/RUB)",
        "关税(%)", "关税(RUB)", "报关手续费(%)", "报关手续费(RUB)",
        "报废税(RUB)", "仓储费(RUB)", "海关费(RUB)", "代理费(RUB)",
        "卢布费用合计", "增值税(%)", "增值税(RUB)",
        "DAP满洲里到岸总价(RUB)", "核验参数(吨位/马力/斗容)"
    ]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
    
    widths = [5, 25, 5, 12, 14, 38, 32, 7, 14, 12, 7, 12, 9, 11, 11, 10, 10, 10, 13, 7, 11, 20, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    
    rn = 2
    for item in inquiries:
        name = item["name"]
        qty = item["qty"]
        if not name or len(name) < 2: continue
        
        candidates, ton, bucket, _ = match_products(name, db, config)
        cat = classify_inquiry(name) or ""
        comp_brand, _ = detect_competitor(name)
        if comp_brand:
            cat = f"竞品({comp_brand})" if not cat else f"{cat}/竞品({comp_brand})"
        
        # No match found
        if not candidates:
            ws.cell(row=rn, column=1, value=item["seq"])
            ws.cell(row=rn, column=2, value=name)
            ws.cell(row=rn, column=3, value=qty)
            ws.cell(row=rn, column=4, value=cat or "非设备")
            reason = "非柳工产品" if is_non_liugong(name) else "未找到匹配"
            ws.cell(row=rn, column=7, value=reason)
            for c in range(1, len(headers)+1):
                cell = ws.cell(row=rn, column=c)
                cell.border = border; cell.alignment = cell_align
                cell.fill = red
            rn += 2; continue
        
        # Show top 4 matches
        nshow = min(len(candidates), 4)
        for ci, cand in enumerate(candidates[:nshow]):
            p = cand["product"]
            cost = calc_dap(p.get("dap_price_cny", 0) or 0, p.get("scrap_tax_rub", 0) or 0, config)
            
            reasons_text = "; ".join(cand["reasons"]) if cand["reasons"] else "综合匹配"
            if ci > 0: reasons_text = f"[备选{ci}] " + reasons_text
            
            verify = ""
            if cand["pton"] > 0: verify += f"吨位≈{cand['pton']}t"
            if p.get("hp", 0) > 0: verify += f"; {p['hp']}hp"
            if p.get("bucket_m3", 0) > 0: verify += f"; {p['bucket_m3']}m³"
            if p.get("engine"): verify += f"; 发动机:{p['engine'][:20]}"
            
            row_data = [
                item["seq"] if ci == 0 else "",
                name if ci == 0 else "",
                qty if ci == 0 else "",
                cat if ci == 0 else "",
                p["model"],
                p.get("desc_ru", p.get("desc_cn", ""))[:200],
                reasons_text,
                f"{cand['score']}分",
                cost["dap_cny"], cost["rate"],
                cost["duty_pct"], cost["duty_rub"],
                cost["proc_pct"], cost["proc_rub"],
                cost["scrap_rub"], cost["warehouse_rub"],
                cost["customs_fee_rub"], cost["agent_fee_rub"],
                cost["fixed_rub"],
                cost["vat_pct"], cost["vat_rub"],
                cost["total_rub"], verify,
            ]
            
            fill = green if ci == 0 else (yellow if ci == 1 else None)
            for col, v in enumerate(row_data, 1):
                cell = ws.cell(row=rn, column=col, value=v)
                cell.border = border; cell.alignment = cell_align
                if fill: cell.fill = fill
            rn += 1
        rn += 1
    
    ws.freeze_panes = "A2"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r'[<>:"/\\|?*]', "_", fname)
    out = os.path.join(OUTPUTS_DIR, f"Report_{safe_name}_{ts}.xlsx")
    wb.save(out)
    return out

# ===== LLM-ENHANCED MATCHING =====

def match_with_llm(name, db, config, api_key=None, model="gpt-4o-mini"):
    """Enhanced matching: LLM parses inquiry -> structured params -> engine matches.
    Falls back to regex parser if no API key or LLM fails."""
    from llm_parser import parse_inquiry_with_llm
    
    parsed = parse_inquiry_with_llm(name, api_key, model)
    
    # If LLM marked as non-LiuGong
    if parsed.get("notes") == "non_liugong" or parsed.get("category") is None:
        return [], 0, 0, {"query": name, "parser": "llm", "parsed": parsed, "rejected": "non_liugong"}
    
    # Use LLM-extracted parameters
    ton = parsed.get("tonnage") or 0
    hp = parsed.get("horsepower") or 0
    bucket = parsed.get("bucket_m3") or 0
    cat = parsed.get("category")
    comp_brand = parsed.get("competitor_brand")
    comp_model = parsed.get("competitor_model")
    
    # If competitor detected by LLM
    if parsed.get("is_competitor") and comp_brand and comp_model:
        comp_ton = ton or 0
        # Check local competitor DB
        m_clean = re.sub(r"[()\-.\/\s]", "", comp_model.upper()) if comp_model else ""
        for key, (brand, t, b) in KNOWN_COMPETITORS.items():
            kc = re.sub(r"[()\-.\/\s]", "", key.upper())
            if kc in m_clean or m_clean in kc:
                if t > 0 and not comp_ton:
                    comp_ton = t
                break
        
        # Try online search
        if not comp_ton:
            online = search_competitor_online(comp_brand, comp_model)
            if online and online.get("tonnage", 0) > 0:
                comp_ton = online["tonnage"]
        
        comp_cat = cat
        candidates = match_competitor_to_liugong(comp_brand, comp_model, comp_ton, comp_cat, db, config)
        if candidates and candidates[0]["score"] > 5:
            verification = {
                "query": name, "parser": "llm" if api_key else "regex",
                "parsed": parsed, "classified_cat": comp_cat,
                "extracted_ton": comp_ton, "competitor": comp_brand,
                "competitor_model": comp_model
            }
            return candidates, comp_ton, bucket, verification
    
    # Normal matching with LLM parameters
    candidates = []
    for p in db:
        pcat = CAT_CN.get(p["category"], p["category"])
        if cat and pcat != cat:
            continue
        
        pton = model_tonnage(p["model"]) or p.get("tonnage", 0)
        score = 0
        reasons = []
        
        if ton > 0 and pton > 0:
            diff_pct = abs(ton - pton) / max(ton, 0.1)
            if ton < 5:
                if diff_pct <= 0.10: score += 50; reasons.append(f"吨位精准匹配({pton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.20: score += 35; reasons.append(f"吨位接近({pton}t,需求{ton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.35: score += 15; reasons.append(f"吨位可参考({pton}t,差{diff_pct:.0%})")
                else: score -= 15; reasons.append(f"吨位不匹配({pton}t,差{diff_pct:.0%})")
            elif ton < 20:
                if diff_pct <= 0.10: score += 50; reasons.append(f"吨位精准匹配({pton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.25: score += 35; reasons.append(f"吨位接近({pton}t,需求{ton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.45: score += 15; reasons.append(f"吨位可参考({pton}t,差{diff_pct:.0%})")
                else: score -= 10; reasons.append(f"吨位不匹配({pton}t,差{diff_pct:.0%})")
            elif ton < 50:
                if diff_pct <= 0.15: score += 50; reasons.append(f"吨位精准匹配({pton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.30: score += 35; reasons.append(f"吨位接近({pton}t,需求{ton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.50: score += 15; reasons.append(f"吨位可参考({pton}t,差{diff_pct:.0%})")
                else: score -= 5
            else:
                if diff_pct <= 0.20: score += 50; reasons.append(f"吨位精准匹配({pton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.40: score += 35; reasons.append(f"吨位接近({pton}t,需求{ton}t,差{diff_pct:.0%})")
                elif diff_pct <= 0.60: score += 15; reasons.append(f"吨位可参考({pton}t,差{diff_pct:.0%})")
                else: score -= 5
        
        pbucket = p.get("bucket_m3", 0)
        if bucket > 0 and pbucket > 0 and cat in ("挖掘机", "装载机"):
            bdiff = abs(bucket - pbucket)
            if bdiff < 0.1: score += 30; reasons.append(f"斗容匹配({pbucket}m3)")
            elif bdiff < 0.5: score += 15; reasons.append(f"斗容接近({pbucket}m3)")
        
        ph = p.get("hp", 0)
        if hp > 0 and ph > 0:
            hdiff = abs(hp - ph) / max(hp, 1)
            if hdiff < 0.1: score += 20; reasons.append(f"马力匹配({ph}hp)")
            elif hdiff < 0.3: score += 10; reasons.append(f"马力接近({ph}hp)")
        
        if p.get("dap_price_cny"): score += 10
        if cat and pcat == cat: score += 5
        
        candidates.append({"product": p, "score": score, "reasons": reasons, "pton": pton})
    
    candidates.sort(key=lambda x: x["score"], reverse=True)
    
    if cat == "压路机" and candidates:
        sub_type = parsed.get("sub_type")
        if sub_type:
            for c in candidates:
                sub = roller_sub(c["product"]["model"])
                if sub and sub_type in sub:
                    c["score"] += 20
                    c["reasons"].append(f"子类匹配({sub})")
            candidates.sort(key=lambda x: x["score"], reverse=True)
    
    verification = {
        "query": name, "parser": "llm" if api_key else "regex",
        "parsed": parsed, "classified_cat": cat,
        "extracted_ton": ton, "extracted_hp": hp, "extracted_bucket": bucket,
        "competitor": comp_brand
    }
    return candidates, ton, bucket, verification

# ===== MAIN =====

def main():
    sys.stdout.reconfigure(encoding="utf-8")
    print("=" * 60)
    print("LiuGong Equipment Matching Engine v19")
    print("=" * 60)
    
    config = load_config()
    db = load_db()
    comp_db = load_competitor_db()
    
    # Auto-import PDF specs
    try:
        from pdf_spec_importer import scan_and_import
        scan_and_import()
        db = load_db()
    except Exception:
        pass
    
    print(f"产品库: {len(db)} 个型号")
    print(f"汇率: {config.get('exchange_rate_rub_cny')} CNY/RUB, 关税: {config.get('customs_duty_rate_pct')}%")
    
    files = [f for f in os.listdir(INQUIRIES_DIR) if f.endswith((".xlsx", ".xls"))]
    if not files:
        print("inquiries/ 文件夹中没有询盘文件")
        return
    
    for fname in files:
        fpath = os.path.join(INQUIRIES_DIR, fname)
        print(f"\n--- {fname} ---")
        inquiries = parse_inquiry_excel(fpath)
        print(f"解析 {len(inquiries)} 条询盘")
        if not inquiries: continue
        
        out = generate_report(inquiries, db, config, fname)
        print(f"报告: {out}")
        
        for item in inquiries:
            cands, _, _, _ = match_products(item["name"], db, config)
            if cands:
                p = cands[0]
                cost = calc_dap(p["product"].get("dap_price_cny", 0) or 0, p["product"].get("scrap_tax_rub", 0) or 0, config)
                status = f"★ {p['product']['model']} ({p['score']}分)"
            else:
                status = "✗ 非柳工产品" if is_non_liugong(item["name"]) else "✗ 未匹配"
            print(f"  {item['name'][:35]:35s} -> {status}")
    
    print("\n完成!")

if __name__ == "__main__":
    main()
